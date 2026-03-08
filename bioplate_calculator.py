"""
BioPlate Calculator - Python/Streamlit Implementation
Replaces the VBS macro functionality for bioplate data analysis

Features:
- Upload Excel file with Layout, Plate1, Plate2, Plate3, Normalization sheets
- Normalize plate values using normalization factors
- Calculate control averages (DMSO = negative, Positive = positive)
- Calculate ratios (Pos/Neg controls)
- Divide all values by negative control average
- Generate final results with AUC, ratio, percentage, average%, and STDEV
- Download processed Excel file with Calculator and FinalResults sheets
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Page configuration
st.set_page_config(
    page_title="BioPlate Calculator",
    page_icon="🧬",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    .stAlert {
        margin-top: 1rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        border-radius: 10px;
        padding: 1rem;
        margin: 0.5rem 0;
    }
    .highlight-yellow {
        background-color: #ffff00;
        padding: 2px 6px;
        border-radius: 3px;
    }
    .highlight-red {
        color: #ff0000;
        font-weight: bold;
    }
</style>
""", unsafe_allow_html=True)

# Title and description
st.title("🧬 BioPlate Calculator")
st.markdown("""
This application processes bioplate assay data by:
1. Reading plate layouts and measurement values
2. Applying normalization factors
3. Calculating control averages (DMSO as negative, Positive controls)
4. Computing ratios and percentages relative to negative controls
5. Generating final results with statistics
""")

# Sidebar for configuration
st.sidebar.header("⚙️ Configuration")
ratio_threshold = st.sidebar.slider(
    "Ratio Highlight Threshold",
    min_value=1.0,
    max_value=2.0,
    value=1.1,
    step=0.05,
    help="Ratios above this value will be highlighted in yellow"
)

control_ratio_threshold = st.sidebar.slider(
    "Control Ratio Threshold",
    min_value=1.0,
    max_value=3.0,
    value=1.5,
    step=0.1,
    help="Pos/Neg control ratios above this value will be highlighted"
)

# File uploader
st.header("📁 Upload Data")
uploaded_file = st.file_uploader(
    "Upload Excel file with sheets: Layout, Plate1, Plate2, Plate3, Normalization",
    type=['xlsx', 'xls'],
    help="The Excel file should contain sheets named Layout, Plate1, Plate2, Plate3, and Normalization"
)


def read_plate_data(excel_file):
    """Read all required sheets from the Excel file."""
    required_sheets = ['Layout', 'Plate1', 'Plate2', 'Plate3', 'Normalization']
    sheets = {}

    try:
        xl = pd.ExcelFile(excel_file)
        available_sheets = xl.sheet_names

        missing_sheets = [s for s in required_sheets if s not in available_sheets]
        if missing_sheets:
            st.error(f"Missing required sheets: {', '.join(missing_sheets)}")
            st.info(f"Available sheets: {', '.join(available_sheets)}")
            return None

        for sheet_name in required_sheets:
            sheets[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

        return sheets
    except Exception as e:
        st.error(f"Error reading Excel file: {str(e)}")
        return None


def extract_well_data(sheets):
    """Extract data from plate layout format to row-based format."""
    layout = sheets['Layout']
    plate1 = sheets['Plate1']
    plate2 = sheets['Plate2']
    plate3 = sheets['Plate3']
    norm = sheets['Normalization']

    # Find dimensions (rows and columns with data)
    last_row = layout.shape[0]
    last_col = layout.shape[1]

    data = []

    # Loop through the plate layout (skip header row/column)
    for i in range(1, last_row):
        for j in range(1, last_col):
            # Get row letter from first column
            row_letter = layout.iloc[i, 0] if pd.notna(layout.iloc[i, 0]) else ""

            # Create well position (e.g., A1, B2)
            well_pos = f"{row_letter}{j}"

            # Get sample ID from layout
            sample_id = layout.iloc[i, j] if pd.notna(layout.iloc[i, j]) else None

            if sample_id is not None and str(sample_id).strip() != "":
                # Get values from each plate and normalization
                try:
                    p1_val = plate1.iloc[i, j] if pd.notna(plate1.iloc[i, j]) else None
                except:
                    p1_val = None

                try:
                    p2_val = plate2.iloc[i, j] if pd.notna(plate2.iloc[i, j]) else None
                except:
                    p2_val = None

                try:
                    p3_val = plate3.iloc[i, j] if pd.notna(plate3.iloc[i, j]) else None
                except:
                    p3_val = None

                try:
                    norm_val = norm.iloc[i, j] if pd.notna(norm.iloc[i, j]) else None
                except:
                    norm_val = None

                data.append({
                    'Well Position': well_pos,
                    'Sample ID': str(sample_id),
                    'Plate1 Value': p1_val,
                    'Plate2 Value': p2_val,
                    'Plate3 Value': p3_val,
                    'Norm Factor': norm_val
                })

    return pd.DataFrame(data)


def calculate_normalized_values(df):
    """Apply normalization factors to plate values."""
    df = df.copy()

    # Calculate normalized values
    for plate in ['Plate1', 'Plate2', 'Plate3']:
        col_name = f'{plate} Value'
        norm_col = f'{plate} Norm'

        df[norm_col] = df.apply(
            lambda row: row[col_name] * row['Norm Factor']
            if pd.notna(row[col_name]) and pd.notna(row['Norm Factor']) and row['Norm Factor'] != 0
            else row[col_name],
            axis=1
        )

    return df


def identify_controls(df):
    """Identify negative (DMSO) and positive controls."""
    df = df.copy()

    df['Is_Neg_Control'] = df['Sample ID'].str.contains('DMSO', case=False, na=False)
    df['Is_Pos_Control'] = df['Sample ID'].str.contains('Positive', case=False, na=False)

    return df


def calculate_control_averages(df):
    """Calculate average values for negative and positive controls."""
    control_stats = {}

    for plate_num in [1, 2, 3]:
        col = f'Plate{plate_num} Norm'

        neg_values = df.loc[df['Is_Neg_Control'], col].dropna()
        pos_values = df.loc[df['Is_Pos_Control'], col].dropna()

        neg_avg = neg_values.mean() if len(neg_values) > 0 else 0
        pos_avg = pos_values.mean() if len(pos_values) > 0 else 0

        control_stats[f'Plate{plate_num}'] = {
            'Neg_Avg': neg_avg,
            'Pos_Avg': pos_avg,
            'Neg_Count': len(neg_values),
            'Pos_Count': len(pos_values),
            'Ratio': pos_avg / neg_avg if neg_avg != 0 else None
        }

    return control_stats


def divide_by_negative_control(df, control_stats):
    """Divide normalized values by negative control averages."""
    df = df.copy()

    for plate_num in [1, 2, 3]:
        norm_col = f'Plate{plate_num} Norm'
        result_col = f'Plate{plate_num} Norm/Neg'
        neg_avg = control_stats[f'Plate{plate_num}']['Neg_Avg']

        df[result_col] = df.apply(
            lambda row: row[norm_col] / neg_avg
            if pd.notna(row[norm_col]) and neg_avg != 0
            else None,
            axis=1
        )

    return df


def generate_final_results(df, control_stats):
    """Generate final results dataframe with AUC, ratio, percentage, avg%, and STDEV."""
    # Filter out controls
    samples = df[~df['Is_Neg_Control'] & ~df['Is_Pos_Control']].copy()

    # Get unique sample IDs
    unique_samples = samples['Sample ID'].unique()

    results = []

    for sample_id in unique_samples:
        sample_data = samples[samples['Sample ID'] == sample_id].iloc[0]

        row = {'Sample ID': sample_id}
        percentages = []

        for plate_num in [1, 2, 3]:
            norm_neg = sample_data.get(f'Plate{plate_num} Norm/Neg')
            neg_avg = control_stats[f'Plate{plate_num}']['Neg_Avg']

            if pd.notna(norm_neg):
                auc = norm_neg * neg_avg
                ratio = norm_neg
                pct = (norm_neg - 1) * 100

                row[f'P{plate_num}_AUC'] = auc
                row[f'P{plate_num}_Ratio'] = ratio
                row[f'P{plate_num}_Pct'] = pct

                percentages.append(pct)
            else:
                row[f'P{plate_num}_AUC'] = None
                row[f'P{plate_num}_Ratio'] = None
                row[f'P{plate_num}_Pct'] = None

        # Calculate average percentage and STDEV
        if percentages:
            row['Avg_Pct'] = np.mean(percentages)
            row['STDEV'] = np.std(percentages, ddof=0) if len(percentages) > 1 else None
        else:
            row['Avg_Pct'] = None
            row['STDEV'] = None

        results.append(row)

    return pd.DataFrame(results)


def create_excel_output(calculator_df, final_results_df, control_stats, sheets, ratio_threshold, control_ratio_threshold):
    """Create formatted Excel output with Calculator and FinalResults sheets."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Copy original sheets
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

        # Write Calculator sheet
        calculator_df.to_excel(writer, sheet_name='Calculator', index=False)

        # Get workbook for formatting
        wb = writer.book
        ws_calc = wb['Calculator']

        # Format Calculator header
        header_fill = PatternFill(start_color='C8C8C8', end_color='C8C8C8', fill_type='solid')
        header_font = Font(bold=True)

        for cell in ws_calc[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Add control statistics to Calculator sheet
        start_col = calculator_df.shape[1] + 3

        ws_calc.cell(row=2, column=start_col, value="Control Averages:").font = Font(bold=True)

        row_offset = 3
        for plate_num in [1, 2, 3]:
            stats = control_stats[f'Plate{plate_num}']

            ws_calc.cell(row=row_offset, column=start_col, value=f"Plate {plate_num} Neg Ctrl Avg:")
            ws_calc.cell(row=row_offset, column=start_col + 1, value=stats['Neg_Avg'])
            row_offset += 1

            ws_calc.cell(row=row_offset, column=start_col, value=f"Plate {plate_num} Pos Ctrl Avg:")
            ws_calc.cell(row=row_offset, column=start_col + 1, value=stats['Pos_Avg'])
            row_offset += 1

            ws_calc.cell(row=row_offset, column=start_col, value=f"Plate {plate_num} Ratio (Pos/Neg):")
            ratio_cell = ws_calc.cell(row=row_offset, column=start_col + 1, value=stats['Ratio'])

            if stats['Ratio'] and stats['Ratio'] > control_ratio_threshold:
                ratio_cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            row_offset += 2

        # Create FinalResults sheet with proper formatting
        ws_final = wb.create_sheet('FinalResults')

        # Write header row 1 values BEFORE merging
        ws_final.cell(row=1, column=1, value='Sample ID')
        ws_final.cell(row=1, column=2, value='plate1')
        ws_final.cell(row=1, column=5, value='plate2')
        ws_final.cell(row=1, column=8, value='plate3')
        ws_final.cell(row=1, column=11, value='average%')
        ws_final.cell(row=1, column=12, value='STDEV')

        # Now merge cells (after writing values to the first cell of each merge range)
        ws_final.merge_cells('B1:D1')
        ws_final.merge_cells('E1:G1')
        ws_final.merge_cells('H1:J1')

        # Apply formatting to header row 1
        for col in range(1, 13):
            cell = ws_final.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # Write header row 2
        headers_row2 = ['', 'AUC', 'ratio', '%', 'AUC', 'ratio', '%', 'AUC', 'ratio', '%', '', '']

        for col, val in enumerate(headers_row2, 1):
            cell = ws_final.cell(row=2, column=col, value=val)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        # Write data rows
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_font = Font(color='FF0000', bold=True)

        for row_idx, row_data in final_results_df.iterrows():
            excel_row = row_idx + 3

            ws_final.cell(row=excel_row, column=1, value=row_data['Sample ID'])

            # Plate 1
            ws_final.cell(row=excel_row, column=2, value=row_data['P1_AUC'])
            ratio_cell_1 = ws_final.cell(row=excel_row, column=3, value=row_data['P1_Ratio'])
            ws_final.cell(row=excel_row, column=4, value=row_data['P1_Pct'])

            # Plate 2
            ws_final.cell(row=excel_row, column=5, value=row_data['P2_AUC'])
            ratio_cell_2 = ws_final.cell(row=excel_row, column=6, value=row_data['P2_Ratio'])
            ws_final.cell(row=excel_row, column=7, value=row_data['P2_Pct'])

            # Plate 3
            ws_final.cell(row=excel_row, column=8, value=row_data['P3_AUC'])
            ratio_cell_3 = ws_final.cell(row=excel_row, column=9, value=row_data['P3_Ratio'])
            ws_final.cell(row=excel_row, column=10, value=row_data['P3_Pct'])

            # Averages
            ws_final.cell(row=excel_row, column=11, value=row_data['Avg_Pct'])
            ws_final.cell(row=excel_row, column=12, value=row_data['STDEV'])

            # Highlight ratios above threshold
            for ratio_cell, ratio_val in [(ratio_cell_1, row_data['P1_Ratio']),
                                           (ratio_cell_2, row_data['P2_Ratio']),
                                           (ratio_cell_3, row_data['P3_Ratio'])]:
                if pd.notna(ratio_val) and ratio_val > ratio_threshold:
                    ratio_cell.fill = yellow_fill

            # Red text for sample ID if all ratios > threshold
            r1, r2, r3 = row_data['P1_Ratio'], row_data['P2_Ratio'], row_data['P3_Ratio']
            if (pd.notna(r1) and pd.notna(r2) and pd.notna(r3) and
                r1 > ratio_threshold and r2 > ratio_threshold and r3 > ratio_threshold):
                ws_final.cell(row=excel_row, column=1).font = red_font

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws_final.iter_rows(min_row=1, max_row=len(final_results_df) + 2, min_col=1, max_col=12):
            for cell in row:
                # Skip merged cells (they don't have border attribute access in some cases)
                try:
                    cell.border = thin_border
                except AttributeError:
                    pass

        # Auto-fit columns using column indices instead of iterating
        from openpyxl.utils import get_column_letter

        for col_idx in range(1, 13):
            max_length = 0
            column_letter = get_column_letter(col_idx)

            for row_idx in range(1, len(final_results_df) + 3):
                try:
                    cell = ws_final.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            ws_final.column_dimensions[column_letter].width = max(max_length + 2, 8)

    output.seek(0)
    return output


# Main processing logic
if uploaded_file is not None:
    st.divider()

    with st.spinner("Reading Excel file..."):
        sheets = read_plate_data(uploaded_file)

    if sheets is not None:
        st.success("✅ File loaded successfully!")

        # Show sheet previews
        with st.expander("📊 Preview Uploaded Sheets", expanded=False):
            tabs = st.tabs(list(sheets.keys()))
            for tab, (name, df) in zip(tabs, sheets.items()):
                with tab:
                    st.dataframe(df.head(10), use_container_width=True)

        st.divider()

        # Process button
        if st.button("🔬 Process Data", type="primary", use_container_width=True):
            with st.spinner("Processing plate data..."):
                # Step 1: Extract well data
                calculator_df = extract_well_data(sheets)
                st.info(f"📍 Extracted {len(calculator_df)} wells with sample data")

                # Step 2: Calculate normalized values
                calculator_df = calculate_normalized_values(calculator_df)

                # Step 3: Identify controls
                calculator_df = identify_controls(calculator_df)

                neg_count = calculator_df['Is_Neg_Control'].sum()
                pos_count = calculator_df['Is_Pos_Control'].sum()
                st.info(f"🎯 Found {neg_count} negative controls (DMSO) and {pos_count} positive controls")

                # Step 4: Calculate control averages
                control_stats = calculate_control_averages(calculator_df)

                # Step 5: Divide by negative control
                calculator_df = divide_by_negative_control(calculator_df, control_stats)

                # Step 6: Generate final results
                final_results_df = generate_final_results(calculator_df, control_stats)

            st.success("✅ Processing complete!")

            st.divider()

            # Display Control Statistics
            st.header("📈 Control Statistics")

            cols = st.columns(3)
            for i, plate_num in enumerate([1, 2, 3]):
                with cols[i]:
                    stats = control_stats[f'Plate{plate_num}']
                    st.subheader(f"Plate {plate_num}")

                    st.metric("Negative Control Avg", f"{stats['Neg_Avg']:.4f}" if stats['Neg_Avg'] else "N/A")
                    st.metric("Positive Control Avg", f"{stats['Pos_Avg']:.4f}" if stats['Pos_Avg'] else "N/A")

                    ratio = stats['Ratio']
                    if ratio:
                        ratio_color = "🟡" if ratio > control_ratio_threshold else "🟢"
                        st.metric(f"Ratio (Pos/Neg) {ratio_color}", f"{ratio:.4f}")
                    else:
                        st.metric("Ratio (Pos/Neg)", "N/A")

            st.divider()

            # Display Calculator Sheet
            st.header("🧮 Calculator Sheet")

            display_cols = ['Well Position', 'Sample ID', 'Plate1 Value', 'Plate2 Value', 'Plate3 Value',
                           'Norm Factor', 'Plate1 Norm', 'Plate2 Norm', 'Plate3 Norm',
                           'Plate1 Norm/Neg', 'Plate2 Norm/Neg', 'Plate3 Norm/Neg']

            display_df = calculator_df[[c for c in display_cols if c in calculator_df.columns]]

            st.dataframe(
                display_df.style.format({
                    col: '{:.4f}' for col in display_df.select_dtypes(include=[np.number]).columns
                }),
                use_container_width=True,
                height=400
            )

            st.divider()

            # Display Final Results
            st.header("📋 Final Results")

            # Highlight function
            def highlight_results(row):
                styles = [''] * len(row)

                # Check ratios for highlighting
                ratio_cols = ['P1_Ratio', 'P2_Ratio', 'P3_Ratio']
                ratio_indices = [final_results_df.columns.get_loc(c) for c in ratio_cols if c in final_results_df.columns]

                all_above_threshold = True
                for idx, col in zip(ratio_indices, ratio_cols):
                    if pd.notna(row[col]) and row[col] > ratio_threshold:
                        styles[idx] = 'background-color: yellow'
                    else:
                        all_above_threshold = False

                # Red text for sample ID if all ratios above threshold
                if all_above_threshold and all(pd.notna(row[c]) for c in ratio_cols if c in row.index):
                    styles[0] = 'color: red; font-weight: bold'

                return styles

            st.dataframe(
                final_results_df.style.apply(highlight_results, axis=1).format({
                    col: '{:.4f}' for col in final_results_df.select_dtypes(include=[np.number]).columns
                }),
                use_container_width=True,
                height=400
            )

            # Summary statistics
            st.subheader("📊 Summary")

            total_samples = len(final_results_df)

            # Count samples with all ratios above threshold
            high_ratio_samples = final_results_df[
                (final_results_df['P1_Ratio'] > ratio_threshold) &
                (final_results_df['P2_Ratio'] > ratio_threshold) &
                (final_results_df['P3_Ratio'] > ratio_threshold)
            ]

            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total Samples", total_samples)
            with col2:
                st.metric("High Ratio Samples (all plates)", len(high_ratio_samples))
            with col3:
                st.metric("Percentage", f"{len(high_ratio_samples)/total_samples*100:.1f}%" if total_samples > 0 else "N/A")

            st.divider()

            # Download button
            st.header("💾 Download Results")

            # Prepare clean dataframe for export (remove boolean columns)
            export_df = calculator_df.drop(columns=['Is_Neg_Control', 'Is_Pos_Control'], errors='ignore')

            excel_output = create_excel_output(
                export_df,
                final_results_df,
                control_stats,
                sheets,
                ratio_threshold,
                control_ratio_threshold
            )

            st.download_button(
                label="📥 Download Processed Excel File",
                data=excel_output,
                file_name="BioPlate_Results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

else:
    # Show example format
    st.header("📖 Expected File Format")

    st.markdown("""
    Your Excel file should contain the following sheets:
    
    | Sheet Name | Description |
    |------------|-------------|
    | **Layout** | Plate layout with sample IDs (row letters in column A, sample IDs in grid) |
    | **Plate1** | Measurement values for Plate 1 (same grid format as Layout) |
    | **Plate2** | Measurement values for Plate 2 (same grid format as Layout) |
    | **Plate3** | Measurement values for Plate 3 (same grid format as Layout) |
    | **Normalization** | Normalization factors (same grid format as Layout) |
    
    ### Control Sample Naming:
    - **Negative controls**: Include "DMSO" in the sample ID (case-insensitive)
    - **Positive controls**: Include "Positive" in the sample ID (case-insensitive)
    
    ### Plate Structure (384-well format):
    - **Rows A & P**: Empty (background/border)
    - **Rows B & O**: Positive controls
    - **Rows C-N**: Samples and DMSO negative controls
    - **Columns 1 & 24**: Empty (border)
    - **Columns 2-23**: Data area
    """)

    # Create realistic sample data for download
    st.subheader("📥 Download Example File")
    st.info("This example file contains realistic data with the same structure as a typical 384-well plate assay.")

    # Generate realistic example data
    sample_output = BytesIO()

    # Define structure
    rows = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']
    cols = list(range(1, 25))

    # Sample IDs
    sample_ids = [
        'BIO-SM-00100-001', 'BIO-SM-00100-002', 'BIO-SM-00200-001', 'BIO-SM-00200-002',
        'BIO-SM-00300-001', 'BIO-SM-00300-002', 'BIO-SM-00400-001', 'BIO-SM-00400-002',
        'BIO-SM-00500-001', 'BIO-SM-00500-002', 'BIO-SM-00600-001', 'BIO-SM-00600-002',
        'BIO-SM-00700-001', 'BIO-SM-00700-002', 'BIO-SM-00800-001', 'BIO-SM-00800-002',
        'BIO-SM-00900-001', 'BIO-SM-00900-002', 'BIO-SM-01000-001', 'BIO-SM-01000-002',
        'BIO-SM-01100-001', 'BIO-SM-01100-002', 'BIO-SM-01200-001', 'BIO-SM-01200-002',
        'BIO-SM-01300-001', 'BIO-SM-01300-002', 'BIO-SM-01400-001', 'BIO-SM-01400-002',
        'BIO-SM-01500-001', 'BIO-SM-01500-002', 'BIO-SM-01600-001', 'BIO-SM-01600-002',
    ]

    # DMSO positions (row, col) - scattered through the plate
    dmso_positions = [
        ('C', 2), ('D', 19), ('E', 8), ('F', 16), ('G', 2), ('H', 5),
        ('I', 10), ('J', 20), ('K', 17), ('L', 13), ('M', 11), ('N', 4)
    ]

    # Create Layout
    layout_data = [[None] + cols]
    sample_idx = 0
    for row_letter in rows:
        row_data = [row_letter]
        for col in cols:
            if row_letter in ['A', 'P']:
                row_data.append(None)
            elif row_letter in ['B', 'O']:
                if col in [1, 24]:
                    row_data.append(None)
                else:
                    row_data.append('Positive')
            else:
                if col in [1, 24]:
                    row_data.append(None)
                elif (row_letter, col) in dmso_positions:
                    row_data.append('DMSO')
                else:
                    row_data.append(sample_ids[sample_idx % len(sample_ids)])
                    sample_idx += 1
        layout_data.append(row_data)

    # Function to create plate data
    def create_plate(base_range, ctrl_range, pos_range, seed):
        np.random.seed(seed)
        plate_data = [[None] + [float(c) for c in cols]]
        for row_letter in rows:
            row_data = [row_letter]
            for col in cols:
                if row_letter in ['A', 'P']:
                    val = np.random.uniform(2500, 3500)
                elif row_letter in ['B', 'O']:
                    if col in [1, 24]:
                        val = np.random.uniform(2500, 3500)
                    else:
                        val = np.random.uniform(*pos_range)
                else:
                    if col in [1, 24]:
                        val = np.random.uniform(2500, 3500)
                    elif (row_letter, col) in dmso_positions:
                        val = np.random.uniform(*ctrl_range)
                    else:
                        val = np.random.uniform(*base_range)
                row_data.append(round(val, 3))
            plate_data.append(row_data)
        return plate_data

    plate1_data = create_plate((8000, 15000), (10000, 12500), (16000, 22000), 42)
    plate2_data = create_plate((7500, 14000), (9500, 11500), (14000, 18000), 43)
    plate3_data = create_plate((7000, 13000), (8500, 10500), (17000, 21000), 44)

    # Create Normalization
    np.random.seed(45)
    norm_data = [[None] + [float(c) for c in cols]]
    for row_letter in rows:
        row_data = [row_letter]
        for col in cols:
            if row_letter in ['A', 'P'] or col in [1, 24]:
                row_data.append(None)
            else:
                row_data.append(round(np.random.uniform(0.88, 1.12), 6))
        norm_data.append(row_data)

    # Write to Excel
    with pd.ExcelWriter(sample_output, engine='openpyxl') as writer:
        pd.DataFrame(layout_data).to_excel(writer, sheet_name='Layout', index=False, header=False)
        pd.DataFrame(plate1_data).to_excel(writer, sheet_name='Plate1', index=False, header=False)
        pd.DataFrame(plate2_data).to_excel(writer, sheet_name='Plate2', index=False, header=False)
        pd.DataFrame(plate3_data).to_excel(writer, sheet_name='Plate3', index=False, header=False)
        pd.DataFrame(norm_data).to_excel(writer, sheet_name='Normalization', index=False, header=False)

    sample_output.seek(0)

    st.download_button(
        label="📥 Download Example File (384-well plate)",
        data=sample_output,
        file_name="BioPlate_Example.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True
    )

# Footer
st.divider()
st.markdown("""
<div style='text-align: center; color: gray;'>
    BioPlate Calculator v1.0 | Python/Streamlit Implementation
    <br> Powered by Biodaat    
</div>
""", unsafe_allow_html=True)
